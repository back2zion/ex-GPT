#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
EX-GPT 지식 관리 시스템
- 문서 파싱 (PDF, HWP, Excel, Word)
- Qdrant 벡터 데이터베이스 관리
- 증분 업데이트 지원
- 관리자 웹 인터페이스
"""

import os
import hashlib
import json
import logging
from datetime import datetime
from pathlib import Path
import uuid
import mimetypes

# 문서 파싱 라이브러리
import PyPDF2
import fitz  # PyMuPDF (더 좋은 PDF 파싱)
import docx
import openpyxl
import pandas as pd
import olefile  # HWP 파일용
import mammoth  # Word 문서 고급 파싱

# 벡터 DB 및 임베딩
from qdrant_client import QdrantClient
from qdrant_client.models import VectorParams, Distance, PointStruct
import sentence_transformers
from sentence_transformers import SentenceTransformer

# 텍스트 처리
import re
from typing import List, Dict, Optional, Tuple
import asyncio
from concurrent.futures import ThreadPoolExecutor

# Flask 관련
from flask import Blueprint, request, jsonify, render_template_string, send_from_directory
from werkzeug.utils import secure_filename

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DocumentParser:
    """다양한 문서 형식 파싱 클래스"""
    
    def __init__(self):
        self.supported_formats = {
            '.pdf': self.parse_pdf,
            '.docx': self.parse_docx,
            '.doc': self.parse_doc,
            '.xlsx': self.parse_excel,
            '.xls': self.parse_excel,
            '.hwp': self.parse_hwp,
            '.txt': self.parse_txt
        }
    
    def parse_pdf(self, file_path: str) -> List[Dict]:
        """PDF 파일 파싱"""
        chunks = []
        try:
            # PyMuPDF 사용 (더 정확한 파싱)
            doc = fitz.open(file_path)
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                text = page.get_text()
                
                if text.strip():
                    chunks.append({
                        'content': text.strip(),
                        'page': page_num + 1,
                        'type': 'pdf_page'
                    })
            doc.close()
            
        except Exception as e:
            logger.error(f"PDF 파싱 오류 ({file_path}): {e}")
            # 폴백: PyPDF2 사용
            try:
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page_num, page in enumerate(pdf_reader.pages):
                        text = page.extract_text()
                        if text.strip():
                            chunks.append({
                                'content': text.strip(),
                                'page': page_num + 1,
                                'type': 'pdf_page'
                            })
            except Exception as e2:
                logger.error(f"PDF 폴백 파싱도 실패: {e2}")
        
        return chunks
    
    def parse_docx(self, file_path: str) -> List[Dict]:
        """DOCX 파일 파싱"""
        chunks = []
        try:
            # mammoth 사용 (더 정확한 파싱)
            with open(file_path, "rb") as docx_file:
                result = mammoth.extract_raw_text(docx_file)
                text = result.value
                
                if text.strip():
                    # 문단별로 분할
                    paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
                    for i, para in enumerate(paragraphs):
                        if len(para) > 50:  # 의미있는 문단만
                            chunks.append({
                                'content': para,
                                'paragraph': i + 1,
                                'type': 'docx_paragraph'
                            })
                
        except Exception as e:
            logger.error(f"DOCX mammoth 파싱 오류, docx 라이브러리 시도: {e}")
            # 폴백: python-docx 사용
            try:
                doc = docx.Document(file_path)
                for i, para in enumerate(doc.paragraphs):
                    if para.text.strip() and len(para.text.strip()) > 50:
                        chunks.append({
                            'content': para.text.strip(),
                            'paragraph': i + 1,
                            'type': 'docx_paragraph'
                        })
            except Exception as e2:
                logger.error(f"DOCX 폴백 파싱도 실패: {e2}")
        
        return chunks
    
    def parse_doc(self, file_path: str) -> List[Dict]:
        """DOC 파일 파싱 (제한적)"""
        # DOC 파일은 복잡하므로 기본적인 처리만
        logger.warning(f"DOC 파일은 제한적으로 지원됩니다: {file_path}")
        return [{'content': f'DOC 파일 {file_path}가 업로드되었습니다. 정확한 파싱을 위해 DOCX로 변환해주세요.', 'type': 'doc_notice'}]
    
    def parse_excel(self, file_path: str) -> List[Dict]:
        """Excel 파일 파싱"""
        chunks = []
        try:
            # openpyxl과 pandas 둘 다 사용
            workbook = openpyxl.load_workbook(file_path)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 데이터 범위 찾기
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                if max_row > 1 and max_col > 1:
                    # 헤더와 데이터 추출
                    headers = []
                    for col in range(1, max_col + 1):
                        header = sheet.cell(row=1, column=col).value
                        headers.append(str(header) if header else f"Column_{col}")
                    
                    # 각 행을 텍스트로 변환
                    for row in range(2, min(max_row + 1, 1000)):  # 최대 1000행
                        row_data = []
                        for col in range(1, max_col + 1):
                            cell_value = sheet.cell(row=row, column=col).value
                            row_data.append(str(cell_value) if cell_value else "")
                        
                        if any(row_data):  # 빈 행이 아닌 경우
                            content = " | ".join([f"{h}: {v}" for h, v in zip(headers, row_data) if v])
                            chunks.append({
                                'content': content,
                                'sheet': sheet_name,
                                'row': row,
                                'type': 'excel_row'
                            })
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Excel 파싱 오류: {e}")
            # 폴백: pandas 사용
            try:
                excel_file = pd.ExcelFile(file_path)
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    for index, row in df.iterrows():
                        content = " | ".join([f"{col}: {val}" for col, val in row.items() if pd.notna(val)])
                        if content:
                            chunks.append({
                                'content': content,
                                'sheet': sheet_name,
                                'row': index + 2,
                                'type': 'excel_row'
                            })
            except Exception as e2:
                logger.error(f"Excel 폴백 파싱도 실패: {e2}")
        
        return chunks
    
    def parse_hwp(self, file_path: str) -> List[Dict]:
        """HWP 파일 파싱"""
        chunks = []
        try:
            # olefile을 사용한 기본적인 HWP 파싱
            f = olefile.OleFileIO(file_path)
            
            # HWP 파일 구조에서 텍스트 추출 시도
            if f.exists('PrvText'):
                encoded_text = f.openstream('PrvText').read()
                try:
                    # 다양한 인코딩 시도
                    for encoding in ['utf-16', 'cp949', 'euc-kr']:
                        try:
                            text = encoded_text.decode(encoding)
                            if text and len(text) > 10:
                                # 문단별로 분할
                                paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
                                for i, para in enumerate(paragraphs):
                                    if len(para) > 20:
                                        chunks.append({
                                            'content': para,
                                            'paragraph': i + 1,
                                            'type': 'hwp_paragraph'
                                        })
                                break
                        except UnicodeDecodeError:
                            continue
                except Exception as e:
                    logger.error(f"HWP 텍스트 디코딩 오류: {e}")
            
            f.close()
            
        except Exception as e:
            logger.error(f"HWP 파싱 오류: {e}")
            chunks.append({
                'content': f'HWP 파일 {os.path.basename(file_path)}가 업로드되었습니다. 정확한 파싱을 위해 다른 형식으로 변환을 권장합니다.',
                'type': 'hwp_notice'
            })
        
        return chunks
    
    def parse_txt(self, file_path: str) -> List[Dict]:
        """TXT 파일 파싱"""
        chunks = []
        try:
            # 다양한 인코딩 시도
            for encoding in ['utf-8', 'cp949', 'euc-kr', 'latin-1']:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        content = f.read()
                    
                    if content.strip():
                        # 문단별로 분할
                        paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
                        for i, para in enumerate(paragraphs):
                            if len(para) > 20:
                                chunks.append({
                                    'content': para,
                                    'paragraph': i + 1,
                                    'type': 'txt_paragraph'
                                })
                    break
                    
                except UnicodeDecodeError:
                    continue
                    
        except Exception as e:
            logger.error(f"TXT 파싱 오류: {e}")
        
        return chunks
    
    def parse_file(self, file_path: str) -> List[Dict]:
        """파일 확장자에 따라 적절한 파서 선택"""
        file_ext = Path(file_path).suffix.lower()
        
        if file_ext in self.supported_formats:
            logger.info(f"파싱 시작: {file_path} ({file_ext})")
            chunks = self.supported_formats[file_ext](file_path)
            logger.info(f"파싱 완료: {len(chunks)}개 청크 생성")
            return chunks
        else:
            logger.warning(f"지원하지 않는 파일 형식: {file_ext}")
            return []

class KnowledgeManager:
    """지식 베이스 관리 클래스"""
    
    def __init__(self, qdrant_host="localhost", qdrant_port=6333, collection_name="documents"):
        self.qdrant_client = QdrantClient(host=qdrant_host, port=qdrant_port)
        self.collection_name = collection_name
        self.parser = DocumentParser()
        
        # 임베딩 모델 초기화
        self.embedding_model = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
        self.vector_size = 384  # MiniLM-L12-v2의 벡터 크기
        
        # 메타데이터 저장 파일
        self.metadata_file = "knowledge_metadata.json"
        self.file_hashes = self.load_metadata()
        
        # 컬렉션 초기화
        self.ensure_collection_exists()
    
    def ensure_collection_exists(self):
        """Qdrant 컬렉션이 존재하는지 확인하고 없으면 생성"""
        try:
            # 컬렉션 목록 확인
            collections = self.qdrant_client.get_collections()
            collection_names = [col.name for col in collections.collections]
            
            if self.collection_name not in collection_names:
                logger.info(f"컬렉션 '{self.collection_name}' 생성 중...")
                
                self.qdrant_client.create_collection(
                    collection_name=self.collection_name,
                    vectors_config=VectorParams(
                        size=self.vector_size,
                        distance=Distance.COSINE
                    )
                )
                logger.info(f"컬렉션 '{self.collection_name}' 생성 완료")
            else:
                logger.info(f"컬렉션 '{self.collection_name}' 존재 확인")
                
        except Exception as e:
            logger.error(f"컬렉션 확인/생성 오류: {e}")
            raise
    
    def load_metadata(self) -> Dict:
        """메타데이터 파일 로드"""
        try:
            if os.path.exists(self.metadata_file):
                with open(self.metadata_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logger.error(f"메타데이터 로드 오류: {e}")
        return {}
    
    def save_metadata(self):
        """메타데이터 파일 저장"""
        try:
            with open(self.metadata_file, 'w', encoding='utf-8') as f:
                json.dump(self.file_hashes, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error(f"메타데이터 저장 오류: {e}")
    
    def get_file_hash(self, file_path: str) -> str:
        """파일의 MD5 해시 계산"""
        hash_md5 = hashlib.md5()
        try:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
        except Exception as e:
            logger.error(f"파일 해시 계산 오류: {e}")
            return ""
        return hash_md5.hexdigest()
    
    def is_file_updated(self, file_path: str) -> bool:
        """파일이 업데이트되었는지 확인"""
        current_hash = self.get_file_hash(file_path)
        filename = os.path.basename(file_path)
        
        if filename in self.file_hashes:
            return self.file_hashes[filename]['hash'] != current_hash
        return True  # 새 파일
    
    def embed_texts(self, texts: List[str]) -> List[List[float]]:
        """텍스트 목록을 벡터로 임베딩"""
        try:
            embeddings = self.embedding_model.encode(texts)
            return embeddings.tolist()
        except Exception as e:
            logger.error(f"임베딩 생성 오류: {e}")
            return [[0.0] * self.vector_size] * len(texts)
    
    def process_file(self, file_path: str, force_update: bool = False) -> Dict:
        """파일 처리 및 벡터 DB 저장"""
        filename = os.path.basename(file_path)
        
        # 파일 업데이트 확인
        if not force_update and not self.is_file_updated(file_path):
            logger.info(f"파일 '{filename}'은 변경되지 않았습니다. 건너뜀.")
            return {
                'success': True,
                'message': f'파일 {filename}은 이미 최신 상태입니다.',
                'chunks_added': 0,
                'skipped': True
            }
        
        try:
            # 기존 파일 데이터 삭제 (증분 업데이트)
            self.delete_file_from_db(filename)
            
            # 파일 파싱
            chunks = self.parser.parse_file(file_path)
            
            if not chunks:
                return {
                    'success': False,
                    'message': f'파일 {filename}에서 텍스트를 추출할 수 없습니다.',
                    'chunks_added': 0
                }
            
            # 임베딩 생성
            texts = [chunk['content'] for chunk in chunks]
            embeddings = self.embed_texts(texts)
            
            # Qdrant에 저장할 포인트 생성
            points = []
            current_time = datetime.now().isoformat()
            file_hash = self.get_file_hash(file_path)
            
            for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
                point = PointStruct(
                    id=str(uuid.uuid4()),
                    vector=embedding,
                    payload={
                        'filename': filename,
                        'content': chunk['content'],
                        'file_hash': file_hash,
                        'chunk_id': i,
                        'total_chunks': len(chunks),
                        'chunk_type': chunk.get('type', 'unknown'),
                        'page': chunk.get('page'),
                        'paragraph': chunk.get('paragraph'),
                        'sheet': chunk.get('sheet'),
                        'row': chunk.get('row'),
                        'upload_time': current_time,
                        'file_size': os.path.getsize(file_path),
                        'file_ext': Path(file_path).suffix.lower()
                    }
                )
                points.append(point)
            
            # Qdrant에 업서트
            self.qdrant_client.upsert(
                collection_name=self.collection_name,
                points=points
            )
            
            # 메타데이터 업데이트
            self.file_hashes[filename] = {
                'hash': file_hash,
                'upload_time': current_time,
                'chunks': len(chunks),
                'file_size': os.path.getsize(file_path),
                'file_ext': Path(file_path).suffix.lower()
            }
            self.save_metadata()
            
            logger.info(f"파일 '{filename}' 처리 완료: {len(chunks)}개 청크")
            
            return {
                'success': True,
                'message': f'파일 {filename}이 성공적으로 처리되었습니다.',
                'chunks_added': len(chunks),
                'skipped': False
            }
            
        except Exception as e:
            logger.error(f"파일 처리 오류 ({filename}): {e}")
            return {
                'success': False,
                'message': f'파일 {filename} 처리 중 오류: {str(e)}',
                'chunks_added': 0
            }
    
    def delete_file_from_db(self, filename: str):
        """특정 파일의 모든 데이터를 DB에서 삭제"""
        try:
            # 파일명으로 필터링하여 삭제
            self.qdrant_client.delete(
                collection_name=self.collection_name,
                points_selector={
                    "filter": {
                        "must": [
                            {
                                "key": "filename",
                                "match": {"value": filename}
                            }
                        ]
                    }
                }
            )
            logger.info(f"파일 '{filename}'의 기존 데이터 삭제 완료")
        except Exception as e:
            logger.warning(f"기존 데이터 삭제 오류 ({filename}): {e}")
    
    def search_documents(self, query: str, limit: int = 5) -> List[Dict]:
        """문서 검색"""
        try:
            # 쿼리 임베딩
            query_vector = self.embedding_model.encode([query])[0].tolist()
            
            # 검색 실행
            search_results = self.qdrant_client.search(
                collection_name=self.collection_name,
                query_vector=query_vector,
                limit=limit,
                with_payload=True
            )
            
            # 결과 포맷팅
            results = []
            for result in search_results:
                results.append({
                    'filename': result.payload.get('filename', 'Unknown'),
                    'content': result.payload.get('content', ''),
                    'score': float(result.score),
                    'page': result.payload.get('page'),
                    'paragraph': result.payload.get('paragraph'),
                    'chunk_type': result.payload.get('chunk_type'),
                    'upload_time': result.payload.get('upload_time')
                })
            
            return results
            
        except Exception as e:
            logger.error(f"문서 검색 오류: {e}")
            return []
    
    def get_collection_stats(self) -> Dict:
        """컬렉션 통계 정보"""
        try:
            info = self.qdrant_client.get_collection(self.collection_name)
            
            # 파일별 통계
            file_stats = {}
            for filename, metadata in self.file_hashes.items():
                file_stats[filename] = {
                    'chunks': metadata['chunks'],
                    'upload_time': metadata['upload_time'],
                    'file_size': metadata['file_size'],
                    'file_ext': metadata['file_ext']
                }
            
            return {
                'total_vectors': info.vectors_count,
                'total_files': len(self.file_hashes),
                'file_details': file_stats,
                'collection_name': self.collection_name
            }
            
        except Exception as e:
            logger.error(f"통계 조회 오류: {e}")
            return {}
    
    def process_directory(self, directory_path: str, force_update: bool = False) -> Dict:
        """디렉토리 내 모든 지원 파일 처리"""
        results = {
            'total_files': 0,
            'processed_files': 0,
            'skipped_files': 0,
            'failed_files': 0,
            'total_chunks': 0,
            'details': []
        }
        
        supported_extensions = {'.pdf', '.docx', '.doc', '.xlsx', '.xls', '.hwp', '.txt'}
        
        try:
            for root, dirs, files in os.walk(directory_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    file_ext = Path(file_path).suffix.lower()
                    
                    if file_ext in supported_extensions:
                        results['total_files'] += 1
                        
                        result = self.process_file(file_path, force_update)
                        results['details'].append({
                            'filename': file,
                            'result': result
                        })
                        
                        if result['success']:
                            if result.get('skipped', False):
                                results['skipped_files'] += 1
                            else:
                                results['processed_files'] += 1
                                results['total_chunks'] += result['chunks_added']
                        else:
                            results['failed_files'] += 1
            
            logger.info(f"디렉토리 처리 완료: {results['processed_files']}/{results['total_files']} 파일")
            
        except Exception as e:
            logger.error(f"디렉토리 처리 오류: {e}")
            results['error'] = str(e)
        
        return results

# Flask Blueprint 생성
knowledge_bp = Blueprint('knowledge', __name__)

# 전역 KnowledgeManager 인스턴스
knowledge_manager = None

def init_knowledge_manager():
    """KnowledgeManager 초기화"""
    global knowledge_manager
    try:
        knowledge_manager = KnowledgeManager()
        logger.info("Knowledge Manager 초기화 완료")
        return True
    except Exception as e:
        logger.error(f"Knowledge Manager 초기화 실패: {e}")
        return False

@knowledge_bp.route('/api/knowledge/upload', methods=['POST'])
def upload_document():
    """문서 업로드 API"""
    if not knowledge_manager:
        return jsonify({'error': 'Knowledge Manager가 초기화되지 않았습니다.'}), 500
    
    try:
        if 'file' not in request.files:
            return jsonify({'error': '파일이 없습니다.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '파일명이 없습니다.'}), 400
        
        # 파일 저장
        filename = secure_filename(file.filename)
        upload_dir = 'uploads'
        os.makedirs(upload_dir, exist_ok=True)
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)
        
        # 파일 처리
        force_update = request.form.get('force_update', 'false').lower() == 'true'
        result = knowledge_manager.process_file(file_path, force_update)
        
        # 임시 파일 삭제
        try:
            os.remove(file_path)
        except:
            pass
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"문서 업로드 오류: {e}")
        return jsonify({'error': f'업로드 중 오류: {str(e)}'}), 500

@knowledge_bp.route('/api/knowledge/stats', methods=['GET'])
def get_knowledge_stats():
    """지식 베이스 통계"""
    if not knowledge_manager:
        return jsonify({'error': 'Knowledge Manager가 초기화되지 않았습니다.'}), 500
    
    try:
        stats = knowledge_manager.get_collection_stats()
        return jsonify(stats)
    except Exception as e:
        logger.error(f"통계 조회 오류: {e}")
        return jsonify({'error': str(e)}), 500

@knowledge_bp.route('/api/knowledge/search', methods=['POST'])
def search_knowledge():
    """지식 베이스 검색"""
    if not knowledge_manager:
        return jsonify({'error': 'Knowledge Manager가 초기화되지 않았습니다.'}), 500
    
    try:
        data = request.json
        query = data.get('query', '')
        limit = data.get('limit', 5)
        
        if not query:
            return jsonify({'error': '검색어가 없습니다.'}), 400
        
        results = knowledge_manager.search_documents(query, limit)
        return jsonify({'results': results})
        
    except Exception as e:
        logger.error(f"검색 오류: {e}")
        return jsonify({'error': str(e)}), 500

@knowledge_bp.route('/api/knowledge/process_directory', methods=['POST'])
def process_directory():
    """디렉토리 일괄 처리"""
    if not knowledge_manager:
        return jsonify({'error': 'Knowledge Manager가 초기화되지 않았습니다.'}), 500
    
    try:
        data = request.json
        directory_path = data.get('directory_path', '')
        force_update = data.get('force_update', False)
        
        if not directory_path or not os.path.exists(directory_path):
            return jsonify({'error': '유효한 디렉토리 경로가 필요합니다.'}), 400
        
        results = knowledge_manager.process_directory(directory_path, force_update)
        return jsonify(results)
        
    except Exception as e:
        logger.error(f"디렉토리 처리 오류: {e}")
        return jsonify({'error': str(e)}), 500

@knowledge_bp.route('/admin/knowledge')
def admin_knowledge():
    """지식 관리 관리자 페이지"""
    # 간단한 HTML 페이지 반환
    html_content = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <title>EX-GPT 지식 관리</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            .card { background: white; padding: 20px; margin: 10px; border: 1px solid #ddd; }
            .btn { padding: 10px 20px; margin: 5px; border: none; cursor: pointer; }
            .btn-primary { background: #007bff; color: white; }
        </style>
    </head>
    <body>
        <h1>EX-GPT 지식 관리 시스템</h1>
        <div class="card">
            <h2>파일 업로드</h2>
            <input type="file" id="fileInput" multiple>
            <button class="btn btn-primary" onclick="uploadFiles()">업로드</button>
        </div>
        <div class="card">
            <h2>검색 테스트</h2>
            <input type="text" id="searchInput" placeholder="검색어 입력">
            <button class="btn btn-primary" onclick="searchDocuments()">검색</button>
            <div id="searchResults"></div>
        </div>
        
        <script>
            async function uploadFiles() {
                const fileInput = document.getElementById('fileInput');
                const files = fileInput.files;
                
                for (let file of files) {
                    const formData = new FormData();
                    formData.append('file', file);
                    
                    try {
                        const response = await fetch('/api/knowledge/upload', {
                            method: 'POST',
                            body: formData
                        });
                        const result = await response.json();
                        console.log('Upload result:', result);
                        alert(`${file.name}: ${result.message}`);
                    } catch (error) {
                        console.error('Upload error:', error);
                        alert(`${file.name}: 업로드 실패`);
                    }
                }
            }
            
            async function searchDocuments() {
                const query = document.getElementById('searchInput').value;
                
                try {
                    const response = await fetch('/api/knowledge/search', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({ query: query, limit: 5 })
                    });
                    const result = await response.json();
                    
                    const resultsDiv = document.getElementById('searchResults');
                    resultsDiv.innerHTML = '<h3>검색 결과:</h3>';
                    
                    result.results.forEach((item, index) => {
                        resultsDiv.innerHTML += `
                            <div style="border: 1px solid #ccc; padding: 10px; margin: 5px;">
                                <h4>${item.filename} (점수: ${item.score.toFixed(3)})</h4>
                                <p>${item.content.substring(0, 200)}...</p>
                            </div>
                        `;
                    });
                } catch (error) {
                    console.error('Search error:', error);
                }
            }
        </script>
    </body>
    </html>
    """
    return html_content

# 사용 예시 및 CLI 도구
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='EX-GPT 지식 관리 시스템')
    parser.add_argument('--init', action='store_true', help='Qdrant 초기화')
    parser.add_argument('--upload', type=str, help='파일 업로드')
    parser.add_argument('--process-dir', type=str, help='디렉토리 일괄 처리')
    parser.add_argument('--search', type=str, help='검색 테스트')
    parser.add_argument('--stats', action='store_true', help='통계 보기')
    parser.add_argument('--force', action='store_true', help='강제 업데이트')
    
    args = parser.parse_args()
    
    # Knowledge Manager 초기화
    try:
        km = KnowledgeManager()
        print("✅ Knowledge Manager 초기화 완료")
    except Exception as e:
        print(f"❌ 초기화 실패: {e}")
        exit(1)
    
    if args.init:
        print("✅ Qdrant 컬렉션 초기화 완료")
    
    elif args.upload:
        result = km.process_file(args.upload, args.force)
        print(f"업로드 결과: {result}")
    
    elif args.process_dir:
        result = km.process_directory(args.process_dir, args.force)
        print(f"일괄 처리 결과: {result}")
    
    elif args.search:
        results = km.search_documents(args.search)
        print(f"검색 결과 ({len(results)}개):")
        for i, result in enumerate(results, 1):
            print(f"{i}. {result['filename']} (유사도: {result['score']:.3f})")
            print(f"   내용: {result['content'][:100]}...")
    
    elif args.stats:
        stats = km.get_collection_stats()
        print("📊 지식 베이스 통계:")
        print(f"  총 벡터 수: {stats.get('total_vectors', 0)}")
        print(f"  총 파일 수: {stats.get('total_files', 0)}")
        print("  파일 목록:")
        for filename, details in stats.get('file_details', {}).items():
            print(f"    - {filename}: {details['chunks']}개 청크, {details['file_size']} bytes")
    
    else:
        parser.print_help()